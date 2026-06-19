"""Create Assets/Data/Gateinfo.xlsx with gate design sheets."""

from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    import subprocess
    import sys

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "Assets" / "Data" / "Gateinfo.xlsx"

HINT_LABELS = {
    "energy_hint_low": "하위권 에너지",
    "energy_hint_mid": "중위권 에너지",
    "energy_hint_high": "상위권 에너지",
    "energy_hint_extreme": "극상위 에너지",
}

GRADES = [
    # grade, energy_min, energy_max, energy_step, auction_type_override
    ("E", 10, 50, 10, ""),
    ("D", 30, 70, 10, ""),
    ("C", 50, 90, 10, ""),
    ("B", 70, 110, 10, ""),
    ("A", 90, 130, 10, ""),
    ("S", 110, 150, 10, "English"),
    ("SS", 130, 170, 10, "English"),
    ("SSS", 150, 190, 10, "English"),
]

# (grade, tier_id, hint_key, energy_min, energy_max, auction_type, display_order)
ENERGY_TIERS = [
    ("E", "low", "energy_hint_low", 10, 20, "Ebay", 1),
    ("E", "mid", "energy_hint_mid", 30, 30, "Ebay", 2),
    ("E", "high", "energy_hint_high", 40, 50, "English", 3),
    ("D", "low", "energy_hint_low", 30, 40, "Ebay", 1),
    ("D", "mid", "energy_hint_mid", 50, 50, "Ebay", 2),
    ("D", "high", "energy_hint_high", 60, 70, "English", 3),
    ("C", "low", "energy_hint_low", 50, 60, "Ebay", 1),
    ("C", "mid", "energy_hint_mid", 70, 70, "Ebay", 2),
    ("C", "high", "energy_hint_high", 80, 90, "English", 3),
    ("B", "low", "energy_hint_low", 70, 80, "Ebay", 1),
    ("B", "mid", "energy_hint_mid", 90, 90, "Ebay", 2),
    ("B", "high", "energy_hint_high", 100, 110, "English", 3),
    ("A", "low", "energy_hint_low", 90, 100, "Ebay", 1),
    ("A", "mid", "energy_hint_mid", 110, 110, "Ebay", 2),
    ("A", "high", "energy_hint_high", 120, 130, "English", 3),
    ("S", "low", "energy_hint_low", 110, 120, "English", 1),
    ("S", "mid", "energy_hint_mid", 130, 130, "English", 2),
    ("S", "high", "energy_hint_high", 140, 150, "English", 3),
    ("SS", "low", "energy_hint_low", 130, 140, "English", 1),
    ("SS", "mid", "energy_hint_mid", 150, 150, "English", 2),
    ("SS", "high", "energy_hint_high", 160, 170, "English", 3),
    ("SSS", "low", "energy_hint_low", 150, 160, "English", 1),
    ("SSS", "mid", "energy_hint_mid", 170, 170, "English", 2),
    ("SSS", "high", "energy_hint_high", 180, 190, "English", 3),
]

# (grade, tier_id, bid_band_min, bid_band_max, reward_gold_min, reward_gold_max,
#  clear_time_sec, ebay_duration_sec, bid_increment, english_round_sec)
ECONOMY = [
    ("E", "low", 50, 80, 100, 200, 1800, 60, 10, 15),
    ("E", "mid", 80, 120, 150, 300, 2400, 75, 15, 15),
    ("E", "high", 120, 180, 250, 450, 3000, 90, 20, 15),
    ("D", "low", 100, 150, 200, 400, 2400, 60, 15, 15),
    ("D", "mid", 150, 220, 350, 550, 3000, 75, 20, 15),
    ("D", "high", 220, 320, 500, 800, 3600, 90, 25, 15),
    ("C", "low", 180, 260, 400, 700, 3000, 60, 25, 15),
    ("C", "mid", 260, 360, 600, 950, 3600, 75, 30, 15),
    ("C", "high", 360, 500, 850, 1300, 4200, 90, 40, 15),
    ("B", "low", 280, 380, 500, 900, 3600, 60, 30, 15),
    ("B", "mid", 380, 500, 750, 1100, 4200, 75, 40, 15),
    ("B", "high", 500, 700, 900, 1400, 4800, 90, 50, 15),
    ("A", "low", 400, 550, 800, 1200, 4200, 60, 40, 15),
    ("A", "mid", 550, 750, 1100, 1600, 4800, 75, 50, 15),
    ("A", "high", 750, 1000, 1500, 2200, 5400, 90, 75, 15),
    ("S", "low", 600, 850, 1200, 1800, 4800, 90, 60, 15),
    ("S", "mid", 850, 1150, 1700, 2500, 5400, 90, 80, 15),
    ("S", "high", 1150, 1600, 2400, 3500, 6000, 90, 100, 15),
    ("SS", "low", 900, 1250, 1800, 2700, 5400, 90, 80, 15),
    ("SS", "mid", 1250, 1700, 2600, 3800, 6000, 90, 100, 15),
    ("SS", "high", 1700, 2300, 3600, 5200, 6600, 90, 125, 15),
    ("SSS", "low", 1400, 1900, 2800, 4200, 6000, 90, 100, 15),
    ("SSS", "mid", 1900, 2600, 4000, 5800, 6600, 90, 125, 15),
    ("SSS", "high", 2600, 3500, 5500, 8000, 7200, 90, 150, 15),
]


def style_header(ws, headers):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def autosize_columns(ws):
    for column_cells in ws.columns:
        length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(length + 2, 40)


def main():
    OUT.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # --- README sheet ---
    readme = wb.active
    readme.title = "README"
    readme["A1"] = "Gateinfo — 던전 게이트 데이터 (Excel에서 편집 후 Unity Import 예정)"
    readme["A1"].font = Font(bold=True, size=12)
    notes = [
        "",
        "시트 구성:",
        "  GateGrades       — 등급별 에너지 범위",
        "  GateEnergyTiers  — 힌트·경매 방식 (에너지 10단위)",
        "  GateAuctionEconomy — 시작가·보상·시간·입찰 단위",
        "",
        "규칙:",
        "  · energy 값은 10 단위만 사용",
        "  · E~A: 하위/중위 = Ebay, 상위 = English",
        "  · S~SSS: auction_type_override = English (전부 영국식)",
        "  · hint_display는 UI에 보여줄 한글 힌트 (유저에게 정확한 수치는 비공개)",
        "",
        "편집 후 Unity에서 Excel → CSV/게임데이터 Import 도구로 적용 (추후 구현)",
    ]
    for i, line in enumerate(notes, start=2):
        readme[f"A{i}"] = line
    readme.column_dimensions["A"].width = 72

    # --- GateGrades ---
    ws_grades = wb.create_sheet("GateGrades")
    grade_headers = [
        "grade",
        "energy_min",
        "energy_max",
        "energy_step",
        "auction_type_override",
        "notes",
    ]
    style_header(ws_grades, grade_headers)
    for row, (grade, emin, emax, step, override) in enumerate(GRADES, start=2):
        note = "S등급 이상: 항상 영국식" if override else ""
        ws_grades.append([grade, emin, emax, step, override, note])
    autosize_columns(ws_grades)

    # --- GateEnergyTiers ---
    ws_tiers = wb.create_sheet("GateEnergyTiers")
    tier_headers = [
        "grade",
        "tier_id",
        "hint_key",
        "hint_display",
        "energy_min",
        "energy_max",
        "auction_type",
        "display_order",
        "notes",
    ]
    style_header(ws_tiers, tier_headers)
    for row_data in ENERGY_TIERS:
        grade, tier_id, hint_key, emin, emax, auction, order = row_data
        hint_display = HINT_LABELS[hint_key]
        note = f"{grade}급 {hint_display}"
        ws_tiers.append(
            [grade, tier_id, hint_key, hint_display, emin, emax, auction, order, note]
        )
    autosize_columns(ws_tiers)

    # --- GateAuctionEconomy ---
    ws_econ = wb.create_sheet("GateAuctionEconomy")
    econ_headers = [
        "grade",
        "tier_id",
        "bid_band_min",
        "bid_band_max",
        "reward_gold_min",
        "reward_gold_max",
        "clear_time_sec",
        "ebay_duration_sec",
        "bid_increment",
        "english_round_sec",
        "notes",
    ]
    style_header(ws_econ, econ_headers)
    for row_data in ECONOMY:
        grade, tier_id, *nums = row_data
        note = "clear_time_sec = 클리어 대기(초), ebay_duration_sec = eBay 경매 시간"
        ws_econ.append([grade, tier_id, *nums, note if row_data == ECONOMY[0] else ""])
    autosize_columns(ws_econ)

    wb.save(OUT)
    print(f"Created: {OUT}")


if __name__ == "__main__":
    main()
